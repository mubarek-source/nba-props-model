# =============================================================================
# ULTORN v5.2 — STEP 2: Model & Predictions
# =============================================================================
# BEFORE RUNNING: fill in Injured? column in picks_today.xlsx first, save it.
#
# Run:  python3 ultorn_model.py
# =============================================================================

import re
import time
import warnings
import numpy as np
import pandas as pd
import openpyxl

from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from nba_api.stats.endpoints import playergamelog, leaguedashteamstats
from nba_api.stats.static import players as nba_players_static

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIG
# =============================================================================
PICKS_FILE     = "picks_today.xlsx"
TRACKER_FILE   = "ultorn_tracker.xlsx"
SEASON         = "2024-25"
ROLLING_GAMES  = 10
MIN_GAMES      = 15
BANKROLL       = 50.0
KELLY_CAP      = 0.10
PROB_THRESHOLD = 0.55
SITE_EDGE_MIN  = 3.0
COMBINED_MIN   = 0.60
API_DELAY      = 0.6

STAT_COL_MAP = {
    "PTS":       "pts",
    "REB":       "reb",
    "AST":       "ast",
    "AST + PTS": "pts_ast",
    "3PM":       "fg3m",
}

TODAY_COL_MAP = {
    "pts":     "today_pts",
    "reb":     "today_reb",
    "ast":     "today_ast",
    "pts_ast": "today_pts_ast",
    "fg3m":    "today_pts",
}

DEF_RANK_MAP = {
    "pts":     "pts_def_rank",
    "reb":     "reb_def_rank",
    "ast":     "ast_def_rank",
    "pts_ast": "pts_def_rank",
    "fg3m":    "pts_def_rank",
}

FEAT_COLS = [
    "target_ewm", "pts_ewm", "reb_ewm", "ast_ewm",
    "min_ewm", "usage_ewm", "target_std", "target_zscore",
    "opp_def_rank", "b2b", "is_home"
]

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DARK   = "1A1A2E"
ORANGE = "FCE4D6"
BLUE   = "DDEEFF"
GOLD   = "FFF9CC"
GRAY1  = "F9F9F9"
GRAY2  = "FFFFFF"

# =============================================================================
# LOAD PICKS FROM EXCEL (after you filled injuries)
# =============================================================================

def load_picks(path=PICKS_FILE):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Today Picks"]
    except Exception as e:
        print(f"  [ERROR] Could not open {path}: {e}")
        print(f"  Run Step 1 first: python3 ultorn_step1_clean.py")
        return []

    picks = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        player_name = str(row[0]).strip()
        team        = str(row[1]).strip() if row[1] else ""
        opp         = str(row[2]).strip() if row[2] else ""
        stat        = str(row[3]).strip() if row[3] else "AST + PTS"
        line        = float(row[4]) if row[4] else None
        site_proj   = float(row[5]) if row[5] else None
        site_diff   = float(row[6]) if row[6] else None
        site_edge   = str(row[7]).strip() if row[7] else ""
        injured     = str(row[8]).strip().upper() if row[8] else ""
        today_pts   = float(row[9]) if row[9] else None
        today_ast   = float(row[10]) if row[10] else None

        if injured in ("Y", "YES", "1", "TRUE"):
            print(f"  [INJURED] Skipping {player_name}")
            continue
        if line is None:
            continue
        if site_edge == "U":
            continue
        if (site_diff or 0) < SITE_EDGE_MIN:
            continue

        today_pts_ast = None
        if today_pts is not None and today_ast is not None:
            today_pts_ast = round(today_pts + today_ast, 1)

        picks.append({
            "player_name":   player_name,
            "team":          team,
            "opp":           opp,
            "stat":          stat,
            "line":          line,
            "site_proj":     site_proj,
            "site_diff":     site_diff,
            "site_edge":     site_edge,
            "today_pts":     today_pts,
            "today_ast":     today_ast,
            "today_pts_ast": today_pts_ast,
        })

    print(f"  {len(picks)} eligible picks loaded from {path}")
    return picks

# =============================================================================
# NBA API
# =============================================================================

_ALL_PLAYERS = None

def get_active_players():
    global _ALL_PLAYERS
    if _ALL_PLAYERS is None:
        _ALL_PLAYERS = nba_players_static.get_active_players()
    return _ALL_PLAYERS


def fuzzy_match_player(raw_name):
    players    = get_active_players()
    parts      = raw_name.strip().split()
    if len(parts) < 2:
        return None
    first_init = parts[0].replace(".", "").upper()
    last       = " ".join(parts[1:]).lower()
    candidates = []
    for p in players:
        fp     = p["full_name"].split()
        p_init = fp[0][0].upper() if fp else ""
        p_last = " ".join(fp[1:]).lower()
        if p_init == first_init and (last in p_last or p_last in last):
            candidates.append(p)
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    exact = [c for c in candidates
             if c["full_name"].split()[-1].lower() == last.split()[-1]]
    return exact[0] if exact else candidates[0]


def fetch_gamelog(player_id):
    try:
        log = playergamelog.PlayerGameLog(
            player_id=player_id, season=SEASON, timeout=30
        )
        df = log.get_data_frames()[0]
        if df.empty:
            return pd.DataFrame()
        df = df.rename(columns={
            "PTS": "pts", "REB": "reb", "AST": "ast", "FG3M": "fg3m",
            "MIN": "min", "FGA": "fga", "GAME_DATE": "game_date",
            "MATCHUP": "matchup"
        })
        df["game_date"]   = pd.to_datetime(df["game_date"])
        df                = df.sort_values("game_date").reset_index(drop=True)
        df["min"]         = pd.to_numeric(df["min"], errors="coerce").fillna(0)
        df["usage_proxy"] = df["fga"] / df["min"].replace(0, np.nan)
        df["days_rest"]   = df["game_date"].diff().dt.days.fillna(3)
        df["b2b"]         = (df["days_rest"] <= 1).astype(int)
        df["is_home"]     = df["matchup"].apply(lambda x: 0 if "@" in str(x) else 1)
        df["pts_ast"]     = df["pts"] + df["ast"]
        if "fg3m" not in df.columns:
            df["fg3m"] = 0
        return df
    except Exception as e:
        print(f"    [WARN] gamelog {player_id}: {e}")
        return pd.DataFrame()


def fetch_def_rankings():
    """Fetch opponent defense rankings. Handles multiple nba_api versions."""
    df = None

    # Try current API parameter style
    for measure_arg in [
        {"measure_type_simple": "Opponent"},
        {"measure_type_player_or_team": "Team"},
        {}
    ]:
        try:
            s = leaguedashteamstats.LeagueDashTeamStats(
                season=SEASON, timeout=30, **measure_arg
            )
            df = s.get_data_frames()[0]
            if not df.empty:
                break
        except Exception:
            continue

    if df is None or df.empty:
        print("  [WARN] Defense rankings unavailable — model still runs without it")
        return {}

    try:
        pts_col = "OPP_PTS" if "OPP_PTS" in df.columns else "PTS"
        reb_col = "OPP_REB" if "OPP_REB" in df.columns else "REB"
        ast_col = "OPP_AST" if "OPP_AST" in df.columns else "AST"
        df["pts_def_rank"] = df[pts_col].rank(ascending=False)
        df["reb_def_rank"] = df[reb_col].rank(ascending=False)
        df["ast_def_rank"] = df[ast_col].rank(ascending=False)
        out = {}
        for _, row in df.iterrows():
            out[row["TEAM_ABBREVIATION"]] = {
                "pts_def_rank": row["pts_def_rank"],
                "reb_def_rank": row["reb_def_rank"],
                "ast_def_rank": row["ast_def_rank"],
            }
        print(f"  {len(out)} teams loaded")
        return out
    except Exception as e:
        print(f"  [WARN] Defense rankings parse failed: {e}")
        return {}

# =============================================================================
# FEATURES + MODEL
# =============================================================================

def build_features(df, stat_col, def_rankings, opp_abbr):
    d = df.copy()
    def ewm(s): return s.ewm(span=ROLLING_GAMES, adjust=False).mean().shift(1)
    d["target_ewm"]    = ewm(d[stat_col])
    d["pts_ewm"]       = ewm(d["pts"])
    d["reb_ewm"]       = ewm(d["reb"])
    d["ast_ewm"]       = ewm(d["ast"])
    d["min_ewm"]       = ewm(d["min"])
    d["usage_ewm"]     = ewm(d["usage_proxy"])
    d["target_std"]    = d[stat_col].shift(1).rolling(ROLLING_GAMES).std()
    d["target_zscore"] = (
        (d[stat_col].shift(1) - d["target_ewm"]) /
        d["target_std"].replace(0, np.nan)
    )
    rank_col           = DEF_RANK_MAP.get(stat_col, "pts_def_rank")
    d["opp_def_rank"]  = def_rankings.get(opp_abbr, {}).get(rank_col, 15.0)
    return d.dropna(subset=FEAT_COLS)


def train_and_predict(df, stat_col, line, def_rankings, opp_abbr):
    if len(df) < MIN_GAMES or stat_col not in df.columns:
        return None
    df = build_features(df, stat_col, def_rankings, opp_abbr)
    if len(df) < 8:
        return None
    df["label"] = (df[stat_col] > line).astype(int)
    df = df.dropna(subset=FEAT_COLS + ["label"])
    if len(df) < 8 or len(df["label"].unique()) < 2:
        return None
    train   = df.iloc[:-1]
    today   = df.iloc[[-1]]
    scaler  = StandardScaler()
    X_tr    = scaler.fit_transform(train[FEAT_COLS].values)
    X_today = scaler.transform(today[FEAT_COLS].values)
    clf     = LogisticRegression(max_iter=500, C=0.5)
    clf.fit(X_tr, train["label"].values)
    return round(clf.predict_proba(X_today)[0][1], 4)


def combined_score(model_prob, site_diff, line, today_proj_val):
    scores, weights = [], []
    if model_prob is not None:
        scores.append(model_prob)
        weights.append(0.50)
    if site_diff is not None and line and line > 0:
        sig = 0.5 + min(site_diff / line, 1.0) * 0.5
        scores.append(sig)
        weights.append(0.25)
    if today_proj_val is not None and line and line > 0:
        diff = today_proj_val - line
        sig  = 0.5 + min(max(diff / line, -1.0), 1.0) * 0.5
        scores.append(sig)
        weights.append(0.25)
    if not scores:
        return None
    return round(sum(s * w for s, w in zip(scores, weights)) / sum(weights), 4)


def kelly_bet(model_prob):
    edge = model_prob - (1 - model_prob)
    if edge <= 0:
        return 0.0
    return round(min(edge, KELLY_CAP) * BANKROLL, 2)

# =============================================================================
# EXCEL — clean, readable formatting
# Columns: Date | Player | Team | Opp | Stat | Line | Site Proj | Site Diff
#          Model % | Combined Score | Bet $ | Hedge $ | Result
# =============================================================================

def _make_tracker(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Picks Log"
    hf  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"].value     = "ULTORN v5.2 — Prediction Tracker"
    ws["A1"].font      = Font(name="Arial", bold=True, color="F0A500", size=13)
    ws["A1"].fill      = PatternFill("solid", start_color=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Legend
    ws.merge_cells("A2:M2")
    ws["A2"].value     = "Orange = both signals agree (best)  |  Blue = model confident  |  Fill Result each night: HIT or MISS"
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 15

    # Headers — simplified, no confusing internal model columns
    headers = [
        "Date", "Player", "Team", "Opp", "Stat", "Line",
        "Site Proj", "Site Diff",
        "Model %", "Combined Score",
        "Bet $", "Hedge $", "Result"
    ]
    widths = [12, 26, 7, 7, 12, 8, 10, 10, 10, 14, 8, 8, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = hf
        c.fill      = PatternFill("solid", start_color=DARK)
        c.alignment = ctr
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 24
    ws.freeze_panes = "A4"
    wb.save(path)


def append_to_tracker(results, path=TRACKER_FILE):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb["Picks Log"]
    except FileNotFoundError:
        _make_tracker(path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Picks Log"]

    nxt   = max(ws.max_row + 1, 4)
    today = datetime.now().strftime("%Y-%m-%d")
    nf    = Font(name="Arial", size=10)
    bf    = Font(name="Arial", size=10, bold=True)
    ctr   = Alignment(horizontal="center", vertical="center")
    lft   = Alignment(horizontal="left",   vertical="center")

    for i, p in enumerate(results):
        row      = nxt + i
        prob     = p.get("model_prob")
        combined = p.get("combined_score")
        diff     = p.get("site_diff") or 0

        if combined and combined >= 0.70:
            bg = ORANGE
        elif prob and prob >= 0.65:
            bg = BLUE
        elif diff >= 8:
            bg = GOLD
        else:
            bg = GRAY1 if i % 2 == 0 else GRAY2

        fill = PatternFill("solid", start_color=bg)

        def wc(col, val, align=ctr, fmt=None, bold=False):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = bf if bold else nf
            c.alignment = align
            c.border    = BORDER
            c.fill      = fill
            if fmt:
                c.number_format = fmt
            ws.row_dimensions[row].height = 18

        wc(1,  today)
        wc(2,  p["player_name"],   align=lft, bold=True)
        wc(3,  p["team"])
        wc(4,  p["opp"])
        wc(5,  p["stat"])
        wc(6,  p["line"],          fmt='0.0')           # e.g. 29.5
        wc(7,  p.get("site_proj"), fmt='0.0')           # e.g. 41.3
        wc(8,  p.get("site_diff"), fmt='0.0')           # e.g. 11.8

        # Model % — shown as whole number percent e.g. "99%" not "0.989"
        prob_pct = round(prob * 100) if prob else None
        wc(9,  prob_pct,           fmt='0"%"')          # e.g. 99%

        # Combined score — shown as percent e.g. "89%" not "0.892"
        comb_pct = round((combined or 0) * 100)
        wc(10, comb_pct,           fmt='0"%"')          # e.g. 89%

        wc(11, p.get("bet_size"),  fmt='"$"0.00')       # e.g. $5.00
        wc(12, p.get("hedge_size"),fmt='"$"0.00')       # e.g. $0.75
        wc(13, "")                                       # HIT or MISS

    wb.save(path)
    print(f"  Saved {len(results)} picks to {path}")

# =============================================================================
# MAIN
# =============================================================================

def run_model():
    print("\n" + "="*58)
    print(f"  ULTORN v5.2 — Step 2: Model")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("="*58)

    print(f"\n[1] Loading picks from {PICKS_FILE}...")
    picks = load_picks(PICKS_FILE)
    if not picks:
        print("\n  Nothing to run. Check your picks file.")
        return []

    print("\n[2] Fetching opponent defense rankings...")
    def_rankings = fetch_def_rankings()

    print("\n[3] Running predictions...")
    results = []

    for pick in picks:
        name      = pick["player_name"]
        stat      = pick["stat"]
        line      = pick["line"]
        opp       = pick["opp"]
        stat_col  = STAT_COL_MAP.get(stat, "pts_ast")
        today_col = TODAY_COL_MAP.get(stat_col, "today_pts_ast")
        today_val = pick.get(today_col)

        player = fuzzy_match_player(name)
        if not player:
            print(f"  [SKIP] No NBA match: {name}")
            continue

        df = fetch_gamelog(player["id"])
        time.sleep(API_DELAY)

        if df.empty:
            print(f"  [SKIP] No game log: {name}")
            continue

        prob = train_and_predict(df, stat_col, line, def_rankings, opp)
        if prob is None:
            print(f"  [SKIP] Not enough data: {name}")
            continue

        comb  = combined_score(prob, pick.get("site_diff"), line, today_val)
        bet   = kelly_bet(prob) if (prob >= PROB_THRESHOLD
                                    and comb is not None
                                    and comb >= COMBINED_MIN) else 0.0
        hedge = round(bet * 0.15, 2)

        if comb and comb >= 0.70:
            flag = "STRONG"
        elif prob >= 0.65:
            flag = "GOOD  "
        else:
            flag = "WEAK  "

        print(f"  [{flag}] {name:<28} {stat:<12} "
              f"line {line:<5}  model {round(prob*100)}%  "
              f"score {round((comb or 0)*100)}%  bet ${bet:.2f}")

        results.append({
            **pick,
            "model_prob":     prob,
            "combined_score": comb,
            "bet_size":       bet,
            "hedge_size":     hedge,
        })

    if not results:
        print("\n  No results generated.")
        return []

    results.sort(key=lambda x: x.get("combined_score") or 0, reverse=True)
    strong = [r for r in results if (r.get("combined_score") or 0) >= COMBINED_MIN]

    # Clean summary
    print(f"\n{'='*58}")
    print(f"  TODAY'S PICKS  ({len(strong)} bets)")
    print(f"{'='*58}")
    print(f"  {'Player':<28} {'Stat':<12} {'Dir':<5} {'Line':<6} {'Score':<8} {'Bet'}")
    print(f"  {'-'*28} {'-'*12} {'-'*5} {'-'*6} {'-'*8} {'-'*6}")
    for r in strong:
        direction = "OVER" if r["model_prob"] >= 0.5 else "UNDER"
        print(f"  {r['player_name']:<28} {r['stat']:<12} {direction:<5} "
              f"{r['line']:<6} {round((r['combined_score'] or 0)*100)}%      "
              f"${r['bet_size']:.2f}")

    total = sum(r["bet_size"] for r in strong)
    print(f"\n  Total at risk: ${total:.2f} of ${BANKROLL:.2f} bankroll")

    print(f"\n  TOP 2 PICKS:")
    for r in results[:2]:
        direction = "OVER" if r["model_prob"] >= 0.5 else "UNDER"
        print(f"  >>> {r['player_name']} — {r['stat']} {direction} {r['line']} "
              f"— {round(r['model_prob']*100)}% confidence — bet ${r['bet_size']:.2f}")

    print(f"\n[4] Saving to {TRACKER_FILE}...")
    append_to_tracker(results, TRACKER_FILE)
    print(f"\n  Done. Open {TRACKER_FILE} and fill in Result tonight.\n")
    return results


if __name__ == "__main__":
    run_model()
